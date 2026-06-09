import os, subprocess, sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK  = "1A1A2E"
GOLD  = "C9A96E"
WHITE = "FFFFFF"
LGREY = "F5F0EA"
GREEN_BG = "E8F5E9"

def fill(h): return PatternFill("solid", fgColor=h)
def bdr():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ── Sheet 1: Brand Accounts ──────────────────────────────────────────────────
ws = wb.active
ws.title = "Brand Accounts"

ws.merge_cells("A1:K1")
ws["A1"] = "PakStyle BD — Brand Account Tracker"
ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=14)
ws["A1"].fill = fill(DARK); ws["A1"].alignment = ctr()
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:K2")
ws["A2"] = "Standard Login:  Email: collectionmoors@gmail.com  |  Name: Mouri  |  Phone: +92-317-3949739  |  Password: Moors@123?"
ws["A2"].font = Font(name="Arial", bold=True, color=DARK, size=9)
ws["A2"].fill = fill("FFF8E1"); ws["A2"].alignment = ctr()
ws.row_dimensions[2].height = 22

ws.merge_cells("A3:K3")
ws["A3"] = "Shipping Address:  ANUM CLASSIC BLACK 'B' ROOM No. M.3, MEZZANINE FLOOR, MAIN SHAHRAH FAISAL, KARACHI  |  Contact: M. Irshad  +92-311-5902323"
ws["A3"].font = Font(name="Arial", bold=False, color="555555", size=9)
ws["A3"].fill = fill("F0F4FF"); ws["A3"].alignment = ctr()
ws.row_dimensions[3].height = 20

headers = ["#","Brand","Website","Signup URL","Platform","Email Used","Password","Account Created?","Notes","Last Order","Account Page"]
col_widths = [4, 20, 26, 44, 12, 28, 16, 16, 22, 14, 34]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fill(DARK); c.alignment = ctr(); c.border = bdr()
ws.row_dimensions[4].height = 28

# Format: (Brand, Website, Signup URL, Platform, Confirmed?)
# Confirmed = True means account already created
brands = [
    # ── CONFIRMED ACCOUNTS ──────────────────────────────────────────────────
    ("Khaadi",           "pk.khaadi.com",           "https://pk.khaadi.com/account/register",             "Shopify", True),
    ("Sapphire",         "pk.sapphireonline.pk",     "https://pk.sapphireonline.pk/account/register",      "Shopify", True),
    ("Limelight",        "limelight.com",            "https://www.limelight.com/account/register",         "Shopify", True),
    ("Farah Talib Aziz", "farahtalibaziz.com.pk",    "https://www.farahtalibaziz.com.pk/account/register", "Shopify", True),
    ("ECS (shopecs)",    "shopecs.com",              "https://www.shopecs.com/account/register",           "Shopify", True),
    ("Bareeze",          "bareezepk.com",            "https://www.bareezepk.com/account/register",         "Shopify", True),
    ("Bonanza Satrangi", "bonanzasatrangi.com",      "https://bonanzasatrangi.com/account/register",       "Shopify", True),
    ("Nishat Linen",     "nishatlinen.com",          "https://nishatlinen.com/account/register",           "Shopify", True),
    ("Rang Ja",          "myrangja.com",             "https://myrangja.com/account/register",              "Shopify", True),
    ("Cross Stitch",     "crossstitch.pk",           "https://www.crossstitch.pk/account/register",        "Shopify", True),
    ("J. Junaid Jamshed","junaidjamshed.com",        "https://www.junaidjamshed.com/account/register",     "Shopify", True),
    ("LAAM",             "laam.pk",                  "https://laam.pk/account/register",                   "Custom",  True),
    ("Ethnicc",          "pk.ethnc.com",             "https://pk.ethnc.com/account/register",              "Shopify", True),
    ("Silai Pret",       "silyaipret.com",           "https://silyaipret.com/account/register",            "Shopify", True),
    # ── ACCOUNTS TO CREATE ──────────────────────────────────────────────────
    ("Maria B",          "mariab.pk",                "https://www.mariab.pk/account/register",             "Shopify", False),
    ("Gul Ahmed",        "gulahmed.com",             "https://www.gulahmed.com/account/register",          "Shopify", False),
    ("Alkaram",          "alkaramstudio.com",        "https://www.alkaramstudio.com/account/register",     "Shopify", False),
    ("Sana Safinaz",     "sanasafinaz.com",          "https://www.sanasafinaz.com/account/register",       "Shopify", False),
    ("Generation",       "generation.com.pk",        "https://www.generation.com.pk/account/register",     "Shopify", False),
    ("Asim Jofa",        "asimjofa.com",             "https://www.asimjofa.com/account/register",          "Shopify", False),
    ("Elan",             "elan.com.pk",              "https://www.elan.com.pk/account/register",           "Shopify", False),
    ("Baroque",          "baroque.pk",               "https://baroque.pk/account/register",                "Shopify", False),
    ("Zara Shahjahan",   "zarashahjahan.com",        "https://www.zarashahjahan.com/account/register",     "Shopify", False),
    ("Rang Rasiya",      "rangrasiya.com",           "https://www.rangrasiya.com/account/register",        "Shopify", False),
    ("Charizma",         "charizma.pk",              "https://charizma.pk/account/register",               "Shopify", False),
    ("Zellbury",         "zellbury.com",             "https://www.zellbury.com/account/register",          "Shopify", False),
    ("Bin Saeed",        "binsaeedstudio.com",       "https://www.binsaeedstudio.com/account/register",    "Shopify", False),
    ("Tawakkal",         "tawakkalonline.com",       "https://www.tawakkalonline.com/account/register",    "Shopify", False),
    ("Barae Khanom",     "baraekhanom.com",          "https://www.baraekhanom.com/account/register",       "Shopify", False),
]

for row_i, (brand, website, signup_url, platform, confirmed) in enumerate(brands, 5):
    bg = GREEN_BG if confirmed else (LGREY if row_i % 2 == 0 else WHITE)
    status = "✓ YES" if confirmed else "No"
    status_color = "1B5E20" if confirmed else "555555"
    vals = [row_i-4, brand, website, signup_url, platform,
            "collectionmoors@gmail.com", "Moors@123?", status, "", "", ""]
    for col_i, val in enumerate(vals, 1):
        c = ws.cell(row=row_i, column=col_i, value=val)
        fc = status_color if col_i == 8 else (DARK if not confirmed else "1B3A1B")
        c.font = Font(name="Arial", bold=(col_i in [2,8]), color=fc, size=9)
        c.fill = fill(bg)
        c.alignment = lft() if col_i not in [1,5,8] else ctr()
        c.border = bdr()
    ws.row_dimensions[row_i].height = 20

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ── Sheet 2: Shipping Address ────────────────────────────────────────────────
ws2 = wb.create_sheet("Shipping Address")
ws2.merge_cells("A1:B1")
ws2["A1"] = "Karachi Logistics Partner — Standard Shipping Address"
ws2["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=13)
ws2["A1"].fill = fill(DARK); ws2["A1"].alignment = ctr()
ws2.row_dimensions[1].height = 30

addr = [
    ("Field", "Value"),
    ("Full Name", "Mouri"),
    ("Company / Store Name", "Collection Moors"),
    ("Full Address", "ANUM CLASSIC BLACK \"B\" ROOM No. M.3, MEZZANINE FLOOR, MAIN SHAHRAH FAISAL, KARACHI"),
    ("If 2 fields required — Line 1", "ANUM CLASSIC BLACK \"B\", ROOM No. M.3, MEZZANINE FLOOR"),
    ("If 2 fields required — Line 2", "MAIN SHAHRAH FAISAL, KARACHI"),
    ("City", "Karachi"),
    ("Province / State", "Sindh"),
    ("Postal Code", "75400"),
    ("Country", "Pakistan"),
    ("Phone", "+92-317-3949739"),
    ("Alternate Contact (M. Irshad)", "+92-311-5902323"),
]

for row_i, (field, value) in enumerate(addr, 2):
    is_hdr = (row_i == 2)
    bg = DARK if is_hdr else (LGREY if row_i % 2 == 0 else WHITE)
    fc = WHITE if is_hdr else DARK
    for col_i, val in enumerate([field, value], 1):
        c = ws2.cell(row=row_i, column=col_i, value=val)
        c.font = Font(name="Arial", bold=(is_hdr or col_i==1), color=fc, size=10)
        c.fill = fill(bg); c.alignment = lft(); c.border = bdr()
    ws2.row_dimensions[row_i].height = 22

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 50

# ── Sheet 3: Order Log ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("Order Log")
ws3.merge_cells("A1:H1")
ws3["A1"] = "PakStyle BD — Brand Order Log"
ws3["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=13)
ws3["A1"].fill = fill(DARK); ws3["A1"].alignment = ctr()
ws3.row_dimensions[1].height = 30

log_hdrs = ["PSB Order ID","Brand","Product URL","Variant / Size","Qty","PKR Price Paid","Brand Order #","Status"]
log_widths = [18, 16, 44, 18, 6, 16, 18, 20]
for i, h in enumerate(log_hdrs, 1):
    c = ws3.cell(row=2, column=i, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fill(DARK); c.alignment = ctr(); c.border = bdr()
ws3.row_dimensions[2].height = 26
for i, w in enumerate(log_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A3"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PakStyleBD_BrandAccounts.xlsx")
wb.save(out)
print("\nDone! File saved to:", out)
input("\nPress Enter to close...")
